#!/usr/bin/env python
# -*- coding: utf-8 -*-
from pathlib import Path
from typing import Any, Dict, Iterator, List, Optional, Protocol, Tuple, Union

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from smorest_sfs.utils.text import camel_to_snake

IDENT_KEY = "<IDENT>"
NAME_KEY = "name"


class IOProtocol(Protocol):
    def read(self) -> bytes:  # pragma: no cover
        ...


class _HierachyParser:
    """
    单excel阅读器

    将文件行转为文件列
    """

    def __init__(
        self,
        filename: Optional[str] = None,
        filepath: Union[str, Path, None] = None,
        filedata: Optional[IOProtocol] = None,
    ):
        self.filepath = Path(filepath) if filepath else None
        self.filename = self.filepath.name if self.filepath else filename
        self.workbook: Workbook = openpyxl.load_workbook(
            filedata if filedata else filepath
        )
        self.relation_sheet: Worksheet
        self.attr_sheet: Worksheet
        self.mapping: Dict[str, Any] = {}
        self.ident: Optional[str] = None
        self.title_list: List[str] = []
        self.relation_list: List[List[Any]] = []

    def _load_workbook(self) -> None:
        self.relation_sheet = self.workbook["relation"]
        self.attr_sheet = self.workbook["attr"]


class _AttrParser:
    attr_sheet: Worksheet
    ident: Optional[str]
    mapping: Dict[str, Any]
    title_list: List[str]

    def _parse_attr_worksheet(self) -> None:
        self.__get_title_list_and_ident()
        self.__check_ident()
        self.__parse_attr_worksheet()

    def __parse_attr_worksheet(self) -> None:
        for data_row in self.attr_sheet.iter_rows(min_row=2, values_only=True):
            item: Dict[str, Any] = dict([*zip(self.title_list, data_row)])
            if self.ident:
                self.mapping[item[self.ident]] = item

    def __get_title_list_and_ident(self) -> None:
        self.title_list = list(self.__iter_raw_title())

    def __iter_raw_title(self) -> Iterator[str]:
        for row in self.attr_sheet.iter_rows(min_row=1, max_row=1, values_only=True):
            for title in self.__parse_titile_row(row):
                yield title

    def __parse_titile_row(self, row: List[str]) -> Iterator[str]:
        for item in row:
            if item:
                yield self.__confirm_ident(item.strip())

    def __check_ident(self) -> None:
        if self.ident is None and NAME_KEY in self.title_list:
            self.ident = NAME_KEY
        elif self.ident is None and NAME_KEY not in self.title_list:
            raise ValueError

    def __confirm_ident(self, ident: str) -> str:
        if IDENT_KEY in ident:
            ident = ident.replace(IDENT_KEY, "")
            self.ident = camel_to_snake(ident)
        return camel_to_snake(ident)


class _RelationParser:
    relation_sheet: Worksheet
    relation_list: List[List[str]]

    @staticmethod
    def __parse_one_col(col: Tuple[Any]) -> List[Any]:
        ret_col = []
        for value in col:
            if value is None:
                break
            ret_col.append(value)
        return ret_col

    def _parse_relation_worksheet(self) -> None:
        for col in self.relation_sheet.iter_cols(values_only=True):
            ret_col = self.__parse_one_col(col)
            self.relation_list.append(ret_col)


class HierachyParser(_HierachyParser, _RelationParser, _AttrParser):
    def _parse_worksheet(self) -> None:
        self._parse_attr_worksheet()
        self._parse_relation_worksheet()

    def parse(self) -> None:
        self._load_workbook()
        self._parse_worksheet()
